#!/usr/bin/env python3
"""
AMOS Daily Brief HTML Exporter Script v0.2

Purpose:
  Export the "Generated Brief" sheet from AMOS Manual Data Workbook v0.2 into a standalone HTML daily brief.

Usage:
  python amos_daily_brief_html_exporter_v0_2.py \
    --workbook amos_manual_data_workbook_v0_2_unified.xlsx \
    --config amos_daily_brief_html_exporter_config_v0_2.json \
    --output amos_daily_brief_exported_from_unified_v0_2.html

Notes:
  - This script is designed for the Codex/engineering stage.
  - It uses openpyxl for a normal local Python environment.
  - In ChatGPT artifact generation, the sample HTML is generated separately.
"""

from __future__ import annotations

import argparse
import json
from collections import Counter
from dataclasses import dataclass
from html import escape
from pathlib import Path
from typing import Any, Iterable, List, Optional


@dataclass
class BriefRow:
    ticker: str
    role: str
    action: str
    freshness: str
    gann: str
    invalidation: str
    note: str


def load_config(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def read_generated_brief_from_xlsx(workbook_path: Path, sheet_name: str) -> list[BriefRow]:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for local workbook export. Install with: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")

    ws = wb[sheet_name]
    header_row_idx: Optional[int] = None
    headers: list[str] = []

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), values_only=True):
        row_values = [safe_str(v).strip() for v in row]
        if "Ticker" in row_values and "Action" in row_values:
            header_row_idx = row[0].row if hasattr(row[0], "row") else None
            headers = row_values
            break

    # openpyxl values_only rows do not expose row index, so do a second explicit scan.
    if header_row_idx is None:
        for idx in range(1, min(ws.max_row, 80) + 1):
            row_values = [safe_str(ws.cell(idx, col).value).strip() for col in range(1, ws.max_column + 1)]
            if "Ticker" in row_values and "Action" in row_values:
                header_row_idx = idx
                headers = row_values
                break

    if header_row_idx is None:
        raise ValueError("Could not find Generated Brief ticker table header.")

    def col_idx(name: str) -> int:
        try:
            return headers.index(name) + 1
        except ValueError:
            return -1

    cols = {
        "ticker": col_idx("Ticker"),
        "role": col_idx("Role"),
        "action": col_idx("Action"),
        "freshness": col_idx("Freshness"),
        "gann": col_idx("Gann"),
        "invalidation": col_idx("Invalidation"),
        "note": col_idx("Note"),
    }

    rows: list[BriefRow] = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        ticker = safe_str(ws.cell(r, cols["ticker"]).value).strip() if cols["ticker"] > 0 else ""
        if not ticker:
            continue
        rows.append(
            BriefRow(
                ticker=ticker,
                role=safe_str(ws.cell(r, cols["role"]).value).strip() if cols["role"] > 0 else "",
                action=safe_str(ws.cell(r, cols["action"]).value).strip() if cols["action"] > 0 else "",
                freshness=safe_str(ws.cell(r, cols["freshness"]).value).strip() if cols["freshness"] > 0 else "",
                gann=safe_str(ws.cell(r, cols["gann"]).value).strip() if cols["gann"] > 0 else "",
                invalidation=safe_str(ws.cell(r, cols["invalidation"]).value).strip() if cols["invalidation"] > 0 else "",
                note=safe_str(ws.cell(r, cols["note"]).value).strip() if cols["note"] > 0 else "",
            )
        )
    return rows


def rows_from_config(config: dict[str, Any]) -> list[BriefRow]:
    return [BriefRow(*r) for r in config.get("fallback_rows", [])]


def build_html(rows: list[BriefRow], config: dict[str, Any]) -> str:
    action_counts = Counter(row.action or "Unknown" for row in rows)
    freshness_counts = Counter(row.freshness or "Unknown" for row in rows)
    gann_counts = Counter(row.gann or "Unknown" for row in rows)
    guardrails = config.get("guardrails", [])

    action_cards = ""
    for action in ["Attack", "Build", "Hold", "Trim Watch", "Exit", "Wait", "Unknown"]:
        count = action_counts.get(action, 0)
        if count == 0:
            continue
        tickers = ", ".join(row.ticker for row in rows if (row.action or "Unknown") == action)
        css = action.lower().replace(" ", "-")
        action_cards += f"""
        <div class="action-card {css}">
          <div class="action-top"><b>{escape(action)}</b><span>{count}</span></div>
          <p>{escape(tickers) if tickers else "—"}</p>
        </div>
        """

    ticker_rows = ""
    for row in rows:
        css = row.action.lower().replace(" ", "-") if row.action else "unknown"
        ticker_rows += f"""
        <tr class="{css}">
          <td><b>{escape(row.ticker)}</b></td>
          <td>{escape(row.role)}</td>
          <td><span class="pill">{escape(row.action)}</span></td>
          <td>{escape(row.freshness)}</td>
          <td>{escape(row.gann)}</td>
          <td>{escape(row.invalidation)}</td>
          <td>{escape(row.note)}</td>
        </tr>
        """

    guardrail_html = "".join(f"<li>{escape(g)}</li>" for g in guardrails)

    summary = f"Hold: {action_counts.get('Hold',0)} | Trim Watch: {action_counts.get('Trim Watch',0)} | Wait: {action_counts.get('Wait',0)} | Missing Data: {freshness_counts.get('Missing',0)} | Gann Live: {gann_counts.get('Gann Live',0)}"

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>AMOS Daily Brief Export v0.2</title>
<style>
:root{{--bg:#05070c;--panel:#0b1020;--line:#1e2a46;--text:#dbe6ff;--muted:#7180a0;--soft:#a9b8d8;--cyan:#23d9ff;--green:#32e68a;--gold:#ffd166;--red:#ff4d6d;--orange:#ff9f43}}
*{{box-sizing:border-box}}
body{{margin:0;background:radial-gradient(circle at 10% 0%,rgba(35,217,255,.08),transparent 30%),radial-gradient(circle at 90% 8%,rgba(255,209,102,.07),transparent 28%),var(--bg);color:var(--text);font-family:Inter,system-ui,-apple-system,"Segoe UI",Arial,"Noto Sans SC",sans-serif;font-size:13px;line-height:1.5}}
.app{{max-width:1520px;margin:0 auto;padding:20px}}.card,.action-card{{background:linear-gradient(180deg,rgba(16,24,45,.95),rgba(8,12,24,.95));border:1px solid var(--line);border-radius:14px}}
.hero{{padding:24px;margin-bottom:14px}}.eyebrow{{font-family:ui-monospace,SFMono-Regular,Menlo,monospace;color:var(--cyan);letter-spacing:.2em;font-size:10px;text-transform:uppercase}}
h1{{font-size:40px;margin:6px 0 8px}}.motto{{font-size:15px;color:var(--gold);font-weight:800}}
.stats{{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin-bottom:14px}}.stat{{padding:16px}}.stat span{{display:block;color:var(--muted);font-family:ui-monospace,SFMono-Regular,Menlo,monospace;font-size:10px;text-transform:uppercase}}.stat b{{display:block;color:var(--green);font-size:28px}}
.actions{{display:grid;grid-template-columns:repeat(6,1fr);gap:10px;margin-bottom:14px}}.action-card{{padding:14px;min-height:130px}}.action-top{{display:flex;justify-content:space-between;gap:8px}}.action-top b{{font-size:16px}}.action-top span{{font-size:28px;color:var(--gold);font-weight:900}}.action-card p{{color:var(--soft)}}
.box{{padding:18px;margin-bottom:14px}}.box h2{{font-size:17px;margin:0 0 10px;text-transform:uppercase;letter-spacing:.1em;color:var(--gold)}}.warning{{border-color:rgba(255,77,109,.55)}}.warning h2{{color:var(--red)}}li{{color:var(--soft);margin:7px 0}}
table{{width:100%;border-collapse:collapse}}th{{color:var(--muted);font-family:ui-monospace,SFMono-Regular,Menlo,monospace;font-size:10px;text-transform:uppercase;text-align:left;padding:9px;border-bottom:1px solid var(--line)}}td{{padding:10px;border-bottom:1px solid rgba(30,42,70,.65);vertical-align:top;color:var(--soft)}}.pill{{border:1px solid var(--line);border-radius:999px;padding:4px 8px;color:var(--gold);font-family:ui-monospace,SFMono-Regular,Menlo,monospace;font-size:10px}}
.trim-watch td{{background:rgba(255,77,109,.05)}}.wait td{{background:rgba(255,209,102,.03)}}
@media(max-width:1100px){{.actions,.stats{{grid-template-columns:1fr 1fr}}}}@media(max-width:700px){{.actions,.stats{{grid-template-columns:1fr}}table{{font-size:11px}}}}
</style>
</head>
<body>
<div class="app">
  <div class="card hero">
    <div class="eyebrow">AMOS · Daily Brief HTML Exporter · v0.2</div>
    <h1>Generated Daily Brief</h1>
    <div class="motto">{escape(summary)}</div>
  </div>

  <div class="stats">
    <div class="card stat"><span>Rows</span><b>{len(rows)}</b></div>
    <div class="card stat"><span>Missing Data</span><b>{freshness_counts.get('Missing',0)}</b></div>
    <div class="card stat"><span>Gann Live</span><b>{gann_counts.get('Gann Live',0)}</b></div>
    <div class="card stat"><span>Trim Watch</span><b>{action_counts.get('Trim Watch',0)}</b></div>
    <div class="card stat"><span>Wait</span><b>{action_counts.get('Wait',0)}</b></div>
  </div>

  <div class="actions">{action_cards}</div>

  <div class="card box warning">
    <h2>Guardrails</h2>
    <ul>{guardrail_html}</ul>
  </div>

  <div class="card box">
    <h2>Ticker Brief Lines</h2>
    <table>
      <thead><tr><th>Ticker</th><th>Role</th><th>Action</th><th>Freshness</th><th>Gann</th><th>Invalidation</th><th>Note</th></tr></thead>
      <tbody>{ticker_rows}</tbody>
    </table>
  </div>
</div>
</body>
</html>"""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=Path("amos_manual_data_workbook_v0_2_unified.xlsx"))
    parser.add_argument("--config", type=Path, default=Path("amos_daily_brief_html_exporter_config_v0_2.json"))
    parser.add_argument("--output", type=Path, default=Path("amos_daily_brief_exported_from_unified_v0_2.html"))
    args = parser.parse_args()

    config = load_config(args.config)

    if args.workbook.exists():
        rows = read_generated_brief_from_xlsx(args.workbook, config.get("source_sheet", "Generated Brief"))
    else:
        rows = rows_from_config(config)

    args.output.write_text(build_html(rows, config), encoding="utf-8")
    print(f"Exported {args.output}")


if __name__ == "__main__":
    main()
